"""
Cape & Cart H1 Growth Campaign — Portfolio Scenario Excel Generator
Dark-theme, insight-driven. Data consistent with dashboard and ebook.
Run: pip install openpyxl && python generate_excel.py
"""
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.chart import BarChart, LineChart, PieChart, Reference
from openpyxl.utils import get_column_letter

wb = Workbook()

# ── PALETTE (dark analytics theme) ──────────────────────────────────────────
DARK      = "F8FAFC"   # page / outer bg
CARD      = "FFFFFF"   # card background
CARD2     = "F1F5F9"   # alt row
BORDER_C  = "CBD5E1"   # subtle borders

TEAL      = "0891B2"   # primary accent
BLUE      = "2563EB"   # Google Ads
GREEN     = "16A34A"   # CM360 / positive
AMBER     = "D97706"   # Bronze / warning
PURPLE    = "7C3AED"   # DV360
RED       = "DC2626"   # YouTube / alert
PINK      = "DB2777"   # decoration

WHITE     = "0F172A"   # bright body text
MUTED     = "334155"   # secondary text
DIM       = "64748B"   # labels

# helpers ────────────────────────────────────────────────────────────────────
def fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def font(bold=False, color=WHITE, size=10, italic=False):
    return Font(bold=bold, color=color, size=size, italic=italic, name="Calibri")

def border(style="thin", color=BORDER_C):
    s = Side(style=style, color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def align(h="left", v="center", wrap=False):
    return Alignment(horizontal=h, vertical=v, wrap_text=wrap)

def set_col_widths(ws, col_widths):
    for col, w in col_widths.items():
        ws.column_dimensions[col].width = w

def header_row(ws, row, values, bg=DARK, fg=TEAL, height=22):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(bold=True, color=fg, size=9)
        c.alignment = align("center")
        c.border = border("thin", BORDER_C)

def data_row(ws, row, values, alt=False, height=18):
    bg = CARD2 if alt else CARD
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(bg)
        c.font = font(color=MUTED, size=9)
        c.alignment = align()
        c.border = border("hair", BORDER_C)

def section_title(ws, row, col, text, accent=TEAL, height=22):
    ws.row_dimensions[row].height = height
    c = ws.cell(row=row, column=col, value=text)
    c.font = font(bold=True, color=accent, size=12)
    c.fill = fill(DARK)

def total_row(ws, row, values, height=20):
    ws.row_dimensions[row].height = height
    for col, val in enumerate(values, 1):
        c = ws.cell(row=row, column=col, value=val)
        c.fill = fill(CARD)
        c.font = font(bold=True, color=WHITE, size=9)
        c.alignment = align("center")
        c.border = border("medium", TEAL)

# ────────────────────────────────────────────────────────────────────────────
# SHEET 1 — COVER
# ────────────────────────────────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Cover"
ws1.sheet_view.showGridLines = False
ws1.sheet_properties.tabColor = TEAL

set_col_widths(ws1, {"A":4,"B":28,"C":18,"D":18,"E":18,"F":18,"G":18})

# Dark header band
for r in range(1, 10):
    ws1.row_dimensions[r].height = 16
    for c_idx in range(1, 8):
        ws1.cell(r, c_idx).fill = fill(DARK)

ws1.merge_cells("B2:G2")
c = ws1["B2"]
c.value = "Cape & Cart H1 Growth Campaign — Portfolio Scenario Pack"
c.font = font(bold=True, color=WHITE, size=22)
c.alignment = align("left")
c.fill = fill(DARK)

ws1.merge_cells("B3:G3")
c = ws1["B3"]
c.value = "Executive decision workbook  ·  1 Jan-30 Jun 2024 data  ·  Anthony Apollis"
c.font = font(color=MUTED, size=11)
c.alignment = align("left")
c.fill = fill(DARK)

ws1.merge_cells("B5:G5")
c = ws1["B5"]
c.value = "Cape & Cart: Cape Town-born ecommerce retailer  ·  Six months  ·  R1.483M paid media spend"
c.font = font(color=TEAL, size=10, italic=True)
c.alignment = align("left")
c.fill = fill(DARK)

# KPI metric boxes
kpi_data = [
    ("8,775",    "Repo Raw Rows",   AMBER),
    ("523K",     "Modelled Rows",   TEAL),
    ("94.0%",    "DQ Pass Rate",    GREEN),
    ("3.8×",     "Blended ROAS",    GREEN),
    ("R1.483M",  "Paid Spend",      PURPLE),
    ("1,886",    "Conversions",     BLUE),
]
for i, (val, lbl, clr) in enumerate(kpi_data):
    col = 2 + i
    r_val, r_lbl = 11, 12
    ws1.row_dimensions[r_val].height = 30
    ws1.row_dimensions[r_lbl].height = 16
    cv = ws1.cell(r_val, col, val)
    cv.font = font(bold=True, color=clr, size=18)
    cv.fill = fill(CARD)
    cv.alignment = align("center")
    cv.border = border("medium", clr)
    cl = ws1.cell(r_lbl, col, lbl)
    cl.font = font(color=DIM, size=8)
    cl.fill = fill(CARD)
    cl.alignment = align("center")
    cl.border = border("hair", BORDER_C)

# Table of contents
ws1.row_dimensions[14].height = 20
ws1.cell(14, 2).value = "CONTENTS"
ws1.cell(14, 2).font = font(bold=True, color=TEAL, size=10)
ws1.cell(14, 2).fill = fill(DARK)
sheets_list = ["Cover","Executive Actions","Product Performance","ML Models","Pipeline Summary","Channel Performance","DQ Report","GA4 Events","Airflow DAG"]
for i, s in enumerate(sheets_list):
    r = 15 + i
    ws1.row_dimensions[r].height = 16
    c = ws1.cell(r, 2, f"Sheet {i+1}: {s}")
    c.font = font(color=BLUE, size=9)
    c.fill = fill(DARK)

# fill remaining bg
for r in range(15, 30):
    ws1.row_dimensions[r].height = 15
    for c_idx in range(1, 8):
        if not ws1.cell(r, c_idx).value:
            ws1.cell(r, c_idx).fill = fill(DARK)

# ────────────────────────────────────────────────────────────────────────────
# SHEET 2 — EXECUTIVE ACTIONS
# ────────────────────────────────────────────────────────────────────────────
ws0 = wb.create_sheet("Executive Actions")
ws0.sheet_view.showGridLines = False
ws0.sheet_properties.tabColor = GREEN
set_col_widths(ws0, {"A":3,"B":26,"C":36,"D":28,"E":30,"F":20})

section_title(ws0, 2, 2, "What I Would Do Monday Morning", GREEN)
ws0.cell(2, 2).fill = fill(DARK)
ws0.cell(3, 2).value = "This workbook is a budget decision brief for Cape & Cart, an anonymised portfolio scenario for a South African ecommerce H1 growth campaign: 1 Jan-30 Jun 2024."
ws0.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws0.cell(3, 2).fill = fill(DARK)

ws0.merge_cells("B4:F4")
ws0.cell(4, 2).value = "Company story: Cape & Cart began as a Cape Town marketplace for curated home, lifestyle, and everyday essentials. By H1 2024 it was scaling nationally, but leadership needed to know whether Search, YouTube, CM360, or DV360 deserved the next rand of budget."
ws0.cell(4, 2).font = font(color=MUTED, size=9, italic=True)
ws0.cell(4, 2).fill = fill(DARK)
ws0.cell(4, 2).alignment = align("left", "center", True)
ws0.row_dimensions[4].height = 42
ws0.merge_cells("B5:F5")
ws0.cell(5, 2).value = "Data scope: the repository contains 8,775 checked-in raw rows across GA4, Google Ads, and YouTube CSVs. The 523,480-row Bronze layer used in the report is a production-scale scenario metric for demonstrating the intended five-source architecture."
ws0.cell(5, 2).font = font(color=MUTED, size=9, italic=True)
ws0.cell(5, 2).fill = fill(DARK)
ws0.cell(5, 2).alignment = align("left", "center", True)
ws0.row_dimensions[5].height = 42

header_row(ws0, 7, ["", "Decision", "Evidence", "Action", "Why it matters", "Owner"], GREEN)
action_rows = [
    ("", "Protect Search scale", "Google Ads ROAS 4.2x on R682K spend", "Increase budget only with marginal ROAS monitoring", "Search is strongest, but may be receiving credit created by YouTube", "Growth + Analytics"),
    ("", "Do not cut YouTube yet", "2.9x last-click ROAS; 1.89M impressions", "Add view-through conversions before budget reduction", "Awareness is under-counted when last click gives Search the conversion", "Marketing Ops"),
    ("", "Recover abandoned carts", "16.8K add-to-cart users did not purchase", "Export GA4 audience to DV360 and bid up on warm users", "This is the highest-intent audience not currently separated", "Lifecycle + Paid Media"),
    ("", "Fix source quality", "31,409 rows quarantined; YouTube date/channel issues", "Correct transfer configuration and monitor reject reasons", "Bad rows are not noise; they are missing attribution evidence", "Data Engineering"),
]
for i, row in enumerate(action_rows):
    r = 8 + i
    data_row(ws0, r, row, alt=(i % 2 == 1), height=36)
    for col in range(2, 7):
        ws0.cell(r, col).alignment = align("left", "center", True)
    ws0.cell(r, 2).font = font(bold=True, color=WHITE, size=9)

section_title(ws0, 15, 2, "Trust Notes", AMBER)
ws0.cell(15, 2).fill = fill(DARK)
header_row(ws0, 17, ["", "Risk", "Control", "Business Meaning"], AMBER)
trust_rows = [
    ("", "Google Ads cost_micros", "Divide by 1,000,000 once in Silver", "Prevents ROAS from being destroyed by unit errors"),
    ("", "Last-click attribution", "Add YouTube view-through metrics", "Stops awareness spend from looking weaker than it is"),
    ("", "Consent Mode modelling", "Track consent-rate drift", "Keeps segment-level ROAS caveated when modelled data rises"),
]
for i, row in enumerate(trust_rows):
    r = 18 + i
    data_row(ws0, r, row, alt=(i % 2 == 1), height=32)
    for col in range(2, 5):
        ws0.cell(r, col).alignment = align("left", "center", True)
    ws0.cell(r, 2).font = font(bold=True, color=AMBER, size=9)

# ────────────────────────────────────────────────────────────────────────────
# ────────────────────────────────────────────────────────────────────────────
# SHEET 3 — PRODUCT PERFORMANCE
# ────────────────────────────────────────────────────────────────────────────
wsP = wb.create_sheet("Product Performance")
wsP.sheet_view.showGridLines = False
wsP.sheet_properties.tabColor = GREEN
set_col_widths(wsP, {"A":3,"B":28,"C":22,"D":18,"E":14,"F":16,"G":12,"H":12})

section_title(wsP, 2, 2, "Product Performance — Modelled Campaign Attribution", GREEN)
wsP.cell(2, 2).fill = fill(DARK)
wsP.cell(3, 2).value = "Product/order data now follows the documented ELT path: dirty Bronze ecommerce exports, 964 quarantined validation failures, clean Silver order lines, and Gold product performance for reporting."
wsP.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
wsP.cell(3, 2).fill = fill(DARK)

header_row(wsP, 5, ["", "Product", "Category", "Lead Channel", "Spend", "Revenue", "ROAS", "Conv"], GREEN)
top_product_rows = [
    ("", "Bamboo Storage Set", "Home Organisation", "Google Ads", 28400, 144800, "5.1x", 312),
    ("", "Loadshedding LED Lamp", "Home Essentials", "Google Ads", 22600, 106200, "4.7x", 241),
    ("", "Cotton Bedding Bundle", "Bedroom", "CM360", 31800, 139900, "4.4x", 198),
    ("", "Airtight Pantry Labels", "Kitchen", "Google Ads", 12800, 55000, "4.3x", 356),
    ("", "Ceramic Dinner Set", "Kitchen", "CM360", 24600, 101900, "4.1x", 144),
    ("", "Reusable Shopper Totes", "Lifestyle", "YouTube", 9600, 38400, "4.0x", 420),
    ("", "Bath Towel Bundle", "Bathroom", "Google Ads", 21400, 81300, "3.8x", 187),
    ("", "Desk Organiser", "Workspace", "DV360", 15800, 58400, "3.7x", 166),
    ("", "Digital Kitchen Scale", "Kitchen", "Google Ads", 13200, 47500, "3.6x", 132),
    ("", "Scented Candle Duo", "Lifestyle", "YouTube", 8800, 29000, "3.3x", 254)
]
for i, row in enumerate(top_product_rows):
    r = 6 + i
    data_row(wsP, r, row, alt=(i % 2 == 1), height=24)
    wsP.cell(r, 2).font = font(bold=True, color=WHITE, size=9)
    wsP.cell(r, 7).font = font(bold=True, color=GREEN, size=9)

section_title(wsP, 18, 2, "Worst 10 Products — Pause, Fix, or Retest", RED)
wsP.cell(18, 2).fill = fill(DARK)
header_row(wsP, 20, ["", "Product", "Category", "Lead Channel", "Spend", "Revenue", "ROAS", "Conv"], RED)
worst_product_rows = [
    ("", "Outdoor Beanbag", "Outdoor", "DV360", 18200, 9100, "0.5x", 12),
    ("", "Glass Terrarium", "Decor", "DV360", 9600, 6700, "0.7x", 18),
    ("", "Luxury Throw Blanket", "Bedroom", "CM360", 15800, 14200, "0.9x", 21),
    ("", "Kids Tee Bundle", "Lifestyle", "YouTube", 7400, 7200, "1.0x", 31),
    ("", "Copper Water Bottle", "Kitchen", "DV360", 11300, 12500, "1.1x", 42),
    ("", "Minimal Wall Clock", "Decor", "CM360", 8900, 10600, "1.2x", 19),
    ("", "Round Pet Bed", "Pet", "DV360", 13400, 17400, "1.3x", 27),
    ("", "Yoga Mat", "Wellness", "YouTube", 7800, 10900, "1.4x", 36),
    ("", "Insulated Travel Mug", "Lifestyle", "DV360", 10100, 15100, "1.5x", 48),
    ("", "Decorative Vase", "Decor", "CM360", 6200, 9900, "1.6x", 22)
]
for i, row in enumerate(worst_product_rows):
    r = 21 + i
    data_row(wsP, r, row, alt=(i % 2 == 1), height=24)
    wsP.cell(r, 2).font = font(bold=True, color=WHITE, size=9)
    wsP.cell(r, 7).font = font(bold=True, color=RED, size=9)

bcP = BarChart()
bcP.type = "bar"
bcP.title = "Top Product Revenue"
bcP.y_axis.title = "Revenue"
bcP.add_data(Reference(wsP, min_col=6, min_row=5, max_row=15), titles_from_data=True)
bcP.set_categories(Reference(wsP, min_col=2, min_row=6, max_row=15))
bcP.series[0].graphicalProperties.solidFill = GREEN
bcP.width = 16; bcP.height = 10
wsP.add_chart(bcP, "J5")

bcW = BarChart()
bcW.type = "bar"
bcW.title = "Worst Product ROAS"
bcW.y_axis.title = "ROAS"
for idx, row in enumerate(worst_product_rows, start=21):
    roas = float(str(row[6]).replace("x", ""))
    wsP.cell(idx, 9, roas)
wsP.cell(20, 9, "ROAS numeric")
bcW.add_data(Reference(wsP, min_col=9, min_row=20, max_row=30), titles_from_data=True)
bcW.set_categories(Reference(wsP, min_col=2, min_row=21, max_row=30))
bcW.series[0].graphicalProperties.solidFill = RED
bcW.width = 16; bcW.height = 10
wsP.add_chart(bcW, "J22")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 4 — ML MODELS
# ────────────────────────────────────────────────────────────────────────────
wsML = wb.create_sheet("ML Models")
wsML.sheet_view.showGridLines = False
wsML.sheet_properties.tabColor = PURPLE
set_col_widths(wsML, {"A":3,"B":28,"C":24,"D":18,"E":14,"F":14,"G":14,"H":14,"I":14})

section_title(wsML, 2, 2, "ML Workbench — GCP Primary, Snowflake Validation", PURPLE)
wsML.cell(2, 2).fill = fill(DARK)
wsML.cell(3, 2).value = "GCP remains the primary platform. Snowflake is used as a validation workbench to prove the curated Silver/Gold data and ML conclusions are portable. Metrics may differ by engine, but features, labels, and windows must match."
wsML.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
wsML.cell(3, 2).fill = fill(DARK)

header_row(wsML, 5, ["", "Use Case", "Best Algorithm", "AUC", "Accuracy", "Precision", "Recall", "F1", "Action"], PURPLE)
ml_rows = [
    ("", "Purchase Propensity", "Logistic Scorecard", 0.4565, 0.8109, 0.1878, 0.0202, 0.0364, "Do not scale yet; add behaviour features"),
    ("", "Churn Risk", "Low-Value Lapse Rule", 0.5038, 0.3642, 0.8202, 0.2879, 0.4262, "Use only for cautious win-back/suppression tests"),
]
for i, row in enumerate(ml_rows):
    r = 6 + i
    data_row(wsML, r, row, alt=(i % 2 == 1), height=30)
    for col in range(2, 10):
        wsML.cell(r, col).alignment = align("left", "center", True)
    wsML.cell(r, 2).font = font(bold=True, color=WHITE, size=9)

section_title(wsML, 10, 2, "ROAS Forecast Algorithm Comparison", TEAL)
wsML.cell(10, 2).fill = fill(DARK)
header_row(wsML, 12, ["", "Algorithm", "MAE", "RMSE", "R2", "Read"], TEAL)
roas_algo_rows = [
    ("", "Trend Projection", 0.0981, 0.1238, 0.9429, "Winner — lowest RMSE and highest R2"),
    ("", "Lag-1 Naive", 0.1300, 0.1378, 0.9292, "Strong simple baseline"),
    ("", "Channel Average", 0.1250, 0.1516, 0.9156, "Stable but less responsive"),
    ("", "Rolling-3 Average", 0.1592, 0.1862, 0.8709, "Too slow to react"),
]
for i, row in enumerate(roas_algo_rows):
    r = 13 + i
    data_row(wsML, r, row, alt=(i % 2 == 1), height=24)
    wsML.cell(r, 2).font = font(bold=True, color=WHITE, size=9)

section_title(wsML, 19, 2, "July Budget Forecast — Best Model", GREEN)
wsML.cell(19, 2).fill = fill(DARK)
header_row(wsML, 21, ["", "Channel", "June Actual ROAS", "July Predicted ROAS", "Action"], GREEN)
forecast_rows = [
    ("", "Google Ads", 4.2, 4.2571, "Scale +10%"),
    ("", "CM360", 3.8, 3.8543, "Hold/test creative"),
    ("", "DV360", 3.4, 3.3114, "Hold/test creative"),
    ("", "YouTube", 2.9, 2.9714, "Reduce -15% and fix audience"),
]
for i, row in enumerate(forecast_rows):
    r = 22 + i
    data_row(wsML, r, row, alt=(i % 2 == 1), height=24)
    wsML.cell(r, 2).font = font(bold=True, color=WHITE, size=9)

bcClass = BarChart()
bcClass.title = "Customer Model AUC"
bcClass.y_axis.title = "AUC"
bcClass.add_data(Reference(wsML, min_col=4, min_row=5, max_row=7), titles_from_data=True)
bcClass.set_categories(Reference(wsML, min_col=2, min_row=6, max_row=7))
bcClass.series[0].graphicalProperties.solidFill = PURPLE
bcClass.width = 13; bcClass.height = 7
wsML.add_chart(bcClass, "K5")

bcErr = BarChart()
bcErr.title = "ROAS RMSE — Lower Is Better"
bcErr.y_axis.title = "RMSE"
bcErr.add_data(Reference(wsML, min_col=4, min_row=12, max_row=16), titles_from_data=True)
bcErr.set_categories(Reference(wsML, min_col=2, min_row=13, max_row=16))
bcErr.series[0].graphicalProperties.solidFill = TEAL
bcErr.width = 13; bcErr.height = 7
wsML.add_chart(bcErr, "K18")

bcML = BarChart()
bcML.title = "July Predicted ROAS"
bcML.y_axis.title = "ROAS"
bcML.add_data(Reference(wsML, min_col=4, min_row=21, max_row=25), titles_from_data=True)
bcML.set_categories(Reference(wsML, min_col=2, min_row=22, max_row=25))
bcML.series[0].graphicalProperties.solidFill = GREEN
bcML.width = 13; bcML.height = 7
wsML.add_chart(bcML, "K31")

# SHEET 5 — PIPELINE SUMMARY# SHEET 5 — PIPELINE SUMMARY
# ────────────────────────────────────────────────────────────────────────────
ws2 = wb.create_sheet("Pipeline Summary")
ws2.sheet_view.showGridLines = False
ws2.sheet_properties.tabColor = AMBER
set_col_widths(ws2, {"A":3,"B":22,"C":20,"D":18,"E":26,"F":22})

section_title(ws2, 2, 2, "Medallion Pipeline Summary", TEAL)
ws2.cell(2, 2).fill = fill(DARK)
ws2.cell(3, 2).value = "Bronze → Silver → Gold  ·  BigQuery Medallion Architecture  ·  H1 2024 campaign data"
ws2.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws2.cell(3, 2).fill = fill(DARK)

header_row(ws2, 5, ["", "Layer", "Dataset", "Row Volume", "Write Mode", "Primary Tool"])
layer_rows = [
    ("", "BRONZE", "marketing_bronze.*", "523,480 modelled rows", "Append (partitioned by date)",  "BQ Data Transfer"),
    ("", "SILVER", "marketing_silver.*", "492,071 rows",     "Incremental merge (unique_key)", "dbt staging models"),
    ("", "GOLD",   "marketing_gold.*",   "fct + rpt tables", "Full table overwrite",           "dbt mart models"),
]
layer_accents = [AMBER, TEAL, GREEN]
for i, row in enumerate(layer_rows):
    r = 6 + i
    data_row(ws2, r, row, alt=(i%2==1))
    ws2.cell(r, 2).font = font(bold=True, color=layer_accents[i], size=9)

section_title(ws2, 11, 2, "Beam Validation — Key Metrics", RED)
ws2.cell(11, 2).fill = fill(DARK)
header_row(ws2, 13, ["", "Metric", "Count", "Pct", "Status"], RED)
dq_summary = [
    ("", "Total Bronze rows ingested", 523480, "100.0%", "Ingested — STRING schema"),
    ("", "Silver — pass validation",   492071,  "94.0%", "Valid rows — typed + deduped"),
    ("", "Quarantined — not dropped",   31409,   "6.0%", "Auditable — reject_reason logged"),
    ("", "Deduplication removed",        3847,   "0.7%", "Silver merge on unique_key"),
    ("", "dbt tests passed",              18,   "100%",  "not_null · unique · range checks"),
]
for i, row in enumerate(dq_summary):
    data_row(ws2, 14+i, row, alt=(i%2==1))
    ws2.cell(14+i, 3).font = font(color=TEAL, bold=True, size=9)

# Pipeline bar chart
for i, (lyr, cnt) in enumerate([("Bronze",523480),("Silver",492071),("Gold (fact)",90)]):
    ws2.cell(22+i, 2, lyr)
    ws2.cell(22+i, 3, cnt)
bc = BarChart()
bc.type = "bar"
bc.title = "Row Volume by Medallion Layer"
bc.y_axis.title = "Rows"
bc.grouping = "clustered"
bc.add_data(Reference(ws2, min_col=3, min_row=22, max_row=24))
bc.set_categories(Reference(ws2, min_col=2, min_row=22, max_row=24))
bc.series[0].graphicalProperties.solidFill = TEAL
bc.width = 16; bc.height = 10
ws2.add_chart(bc, "E5")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 3 — CHANNEL PERFORMANCE
# ────────────────────────────────────────────────────────────────────────────
ws3 = wb.create_sheet("Channel Performance")
ws3.sheet_view.showGridLines = False
ws3.sheet_properties.tabColor = BLUE
set_col_widths(ws3, {"A":3,"B":24,"C":14,"D":12,"E":10,"F":16,"G":14,"H":10,"I":12})

section_title(ws3, 2, 2, "Cross-Channel Performance — Gold Fact Table", TEAL)
ws3.cell(2, 2).fill = fill(DARK)
ws3.cell(3, 2).value = "Source: marketing_gold.fct_cross_channel_performance  ·  1 Jan-30 Jun 2024  ·  Blended ROAS 3.8×"
ws3.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws3.cell(3, 2).fill = fill(DARK)

header_row(ws3, 5, ["","Channel","Impressions","Clicks","CTR","Spend (ZAR)","Conversions","ROAS","Conv Rate"])
ch_data = [
    ("","Google Ads — Search",   3850000, 77000,  "2.0%", 682000, 924, 4.2, "1.20%"),
    ("","Campaign Manager 360",  2140000, 28620,  "1.3%", 341000, 411, 3.8, "1.44%"),
    ("","DV360 — Display",       4210000, 42100,  "1.0%", 278000, 337, 3.4, "0.80%"),
    ("","YouTube Analytics",     1890000, 37800,  "2.0%", 182000, 214, 2.9, "0.57%"),
]
ch_accents = [BLUE, GREEN, PURPLE, RED]
for i, row in enumerate(ch_data):
    r = 6 + i
    data_row(ws3, r, row, alt=(i%2==1))
    # channel name colored
    ws3.cell(r, 2).font = font(bold=True, color=ch_accents[i], size=9)
    # ROAS colored by performance
    roas_val = row[7]
    roas_color = GREEN if roas_val >= 4.0 else (TEAL if roas_val >= 3.5 else (AMBER if roas_val >= 3.0 else RED))
    ws3.cell(r, 8).font = font(bold=True, color=roas_color, size=9)

total_row(ws3, 10, ["","TOTAL PAID","12,090,000","185,520","1.5%","R1,483,000","1,886","3.8×","1.02%"])

# Insight note
ws3.row_dimensions[12].height = 16
ws3.cell(12, 2).value = "INSIGHT: Google Ads delivers 49% of conversions on 46% of budget. DV360 delivers 17.9% of conversions on 18.8% of budget — efficiency gap of 0.9pts on R278K spend."
ws3.cell(12, 2).font = font(color=AMBER, size=8, italic=True)
ws3.cell(12, 2).fill = fill(DARK)

# ROAS chart
for i, (ch, roas) in enumerate([("Google Ads",4.2),("CM360",3.8),("DV360",3.4),("YouTube",2.9)]):
    ws3.cell(15+i, 2, ch)
    ws3.cell(15+i, 3, roas)
bc2 = BarChart()
bc2.type = "bar"
bc2.title = "ROAS by Channel vs 3.0x Target"
bc2.y_axis.title = "ROAS"
bc2.y_axis.scaling.min = 0
bc2.y_axis.scaling.max = 5
bc2.add_data(Reference(ws3, min_col=3, min_row=15, max_row=18))
bc2.set_categories(Reference(ws3, min_col=2, min_row=15, max_row=18))
bc2.series[0].graphicalProperties.solidFill = BLUE
bc2.width = 16; bc2.height = 10
ws3.add_chart(bc2, "E13")

# Impressions pie
for i, (ch, imp) in enumerate([("Google Ads",3850000),("CM360",2140000),("DV360",4210000),("YouTube",1890000)]):
    ws3.cell(22+i, 2, ch)
    ws3.cell(22+i, 3, imp)
pc = PieChart()
pc.title = "Impressions Share by Channel"
pc.add_data(Reference(ws3, min_col=3, min_row=22, max_row=25))
pc.set_categories(Reference(ws3, min_col=2, min_row=22, max_row=25))
pc.width = 16; pc.height = 10
ws3.add_chart(pc, "E25")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 4 — DQ REPORT
# ────────────────────────────────────────────────────────────────────────────
ws4 = wb.create_sheet("DQ Report")
ws4.sheet_view.showGridLines = False
ws4.sheet_properties.tabColor = RED
set_col_widths(ws4, {"A":3,"B":26,"C":16,"D":14,"E":14,"F":14})

section_title(ws4, 2, 2, "Data Quality Report — Apache Beam Validation", RED)
ws4.cell(2, 2).fill = fill(DARK)
ws4.cell(3, 2).value = "ValidateAndTagRow DoFn  ·  Cloud Dataflow  ·  H1 2024 campaign data  ·  6.0% quarantine rate"
ws4.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws4.cell(3, 2).fill = fill(DARK)

header_row(ws4, 5, ["","Source","Total Rows","Passed","Quarantined","Pass Rate"], RED)
dq_by_source = [
    ("","GA4 Events",          148920, 142161,  6759, "95.5%"),
    ("","Google Ads",           89340,  83024,  6316, "92.9%"),
    ("","Campaign Manager 360", 76850,  73490,  3360, "95.6%"),
    ("","DV360",               112430, 105628,  6802, "94.0%"),
    ("","YouTube Analytics",    95940,  87768,  8172, "91.5%"),
]
for i, row in enumerate(dq_by_source):
    r = 6 + i
    data_row(ws4, r, row, alt=(i%2==1))
    rate_str = row[5]
    rate_val = float(rate_str.replace("%",""))
    pass_color = GREEN if rate_val >= 95 else (AMBER if rate_val >= 92 else RED)
    ws4.cell(r, 6).font = font(bold=True, color=pass_color, size=9)

total_row(ws4, 11, ["","TOTAL","523,480","492,071","31,409","94.0%"])

# Issues table
section_title(ws4, 14, 2, "Quarantine Issue Breakdown", AMBER)
ws4.cell(14, 2).fill = fill(DARK)
set_col_widths(ws4, {"B":26,"C":16,"D":20,"E":36})
header_row(ws4, 16, ["","DQ Issue","Row Count","Affected Sources","Root Cause & Fix"], AMBER)
dq_issues = [
    ("","YouTube date format mismatch", 3400, "YouTube",              "DD/MM/YYYY vs YYYY-MM-DD — fix at BQ Transfer config"),
    ("","Missing required field",       8310, "Google Ads · YouTube", "NULL campaign_id or impressions — flagged in quarantine"),
    ("","Negative impressions",         6802, "DV360",                "DV360 billing adjustments — quarantined, not dropped"),
    ("","Missing / invalid YouTube metadata", 6138, "YouTube",              "Incomplete channel/video metadata — quarantined or imputed where recoverable"),
    ("","Currency trailing whitespace", 1840, "Google Ads",           "'ZAR ' → fixed with TRIM(UPPER()) in Beam DoFn"),
    ("","NULL geo_id",                  6759, "GA4",                  "Consent declined / VPN users — kept NULL, not imputed"),
]
for i, row in enumerate(dq_issues):
    data_row(ws4, 17+i, row, alt=(i%2==1))
    ws4.cell(17+i, 3).font = font(color=AMBER, bold=True, size=9)

# Quarantine pie
for i, (s, q) in enumerate([("GA4",6759),("Google Ads",6316),("CM360",3360),("DV360",6802),("YouTube",8172)]):
    ws4.cell(27+i, 2, s)
    ws4.cell(27+i, 3, q)
pc2 = PieChart()
pc2.title = "Quarantined Rows by Source"
pc2.add_data(Reference(ws4, min_col=3, min_row=27, max_row=31))
pc2.set_categories(Reference(ws4, min_col=2, min_row=27, max_row=31))
pc2.width = 16; pc2.height = 10
ws4.add_chart(pc2, "G5")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 5 — GA4 EVENTS
# ────────────────────────────────────────────────────────────────────────────
ws5 = wb.create_sheet("GA4 Events")
ws5.sheet_view.showGridLines = False
ws5.sheet_properties.tabColor = GREEN
set_col_widths(ws5, {"A":3,"B":22,"C":16,"D":14,"E":16,"F":16})

section_title(ws5, 2, 2, "GA4 Event Summary — BigQuery Export", GREEN)
ws5.cell(2, 2).fill = fill(DARK)
ws5.cell(3, 2).value = "analytics_*.events_*  ·  1 Jan-30 Jun 2024  ·  South African ecommerce  ·  Organic CVR 2.5%"
ws5.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws5.cell(3, 2).fill = fill(DARK)

header_row(ws5, 5, ["","Event Name","Event Count","Unique Users","Avg Value (ZAR)","Top Device"], GREEN)
ga4_events = [
    ("","session_start",    480000, 142000,   0.00, "Desktop"),
    ("","page_view",       1240000, 138500,   0.00, "Desktop"),
    ("","search",            38400,  22100,   0.00, "Mobile"),
    ("","add_to_cart",       61200,  29400,  52.40, "Desktop"),
    ("","begin_checkout",    28900,  17800,  71.20, "Desktop"),
    ("","purchase",          12100,   9800,  89.60, "Desktop"),
]
for i, row in enumerate(ga4_events):
    r = 6 + i
    data_row(ws5, r, row, alt=(i%2==1))
    ws5.cell(r, 3).font = font(color=TEAL, bold=True, size=9)

# Funnel
section_title(ws5, 14, 2, "Ecommerce Funnel — Session to Purchase", TEAL)
ws5.cell(14, 2).fill = fill(DARK)
header_row(ws5, 16, ["","Funnel Step","Count","Drop-off","Cumulative Rate"], TEAL)
funnel = [
    ("","Sessions",          480000, "—",     "100.0%"),
    ("","Product views",     214000, "266,000","44.6%"),
    ("","Add to cart",        61200, "152,800","12.8%"),
    ("","Begin checkout",     28900,  "32,300", "6.0%"),
    ("","Purchase",           12100,  "16,800", "2.5%"),
]
for i, row in enumerate(funnel):
    r = 17 + i
    data_row(ws5, r, row, alt=(i%2==1))
    rate_str = row[4]
    rate_num = float(rate_str.replace("%","")) if rate_str != "—" else 100
    c_color = GREEN if rate_num > 10 else (AMBER if rate_num > 4 else RED)
    ws5.cell(r, 5).font = font(bold=True, color=c_color, size=9)

ws5.row_dimensions[23].height = 16
ws5.cell(23, 2).value = "INSIGHT: 58% of users who begin checkout do not purchase. Checkout abandoners are the highest-intent retargeting segment — export this GA4 audience to DV360."
ws5.cell(23, 2).font = font(color=AMBER, size=8, italic=True)
ws5.cell(23, 2).fill = fill(DARK)

# Monthly sessions line chart
for i, (m, s) in enumerate([("Jan",62000),("Feb",74000),("Mar",81000),("Apr",88000),("May",94000),("Jun",81000)]):
    ws5.cell(27+i, 2, m)
    ws5.cell(27+i, 3, s)
lc = LineChart()
lc.title = "Monthly Sessions — GA4 H1 2024 Campaign"
lc.y_axis.title = "Sessions"
lc.add_data(Reference(ws5, min_col=3, min_row=27, max_row=32))
lc.set_categories(Reference(ws5, min_col=2, min_row=27, max_row=32))
lc.series[0].graphicalProperties.line.solidFill = GREEN
lc.series[0].graphicalProperties.line.width = 20000
lc.width = 18; lc.height = 10
ws5.add_chart(lc, "E5")

# ────────────────────────────────────────────────────────────────────────────
# SHEET 6 — AIRFLOW DAG
# ────────────────────────────────────────────────────────────────────────────
ws6 = wb.create_sheet("Airflow DAG")
ws6.sheet_view.showGridLines = False
ws6.sheet_properties.tabColor = PURPLE
set_col_widths(ws6, {"A":3,"B":26,"C":20,"D":20,"E":20,"F":16})

section_title(ws6, 2, 2, "Airflow DAG — gcp_marketing_analytics_pipeline", PURPLE)
ws6.cell(2, 2).fill = fill(DARK)
ws6.cell(3, 2).value = "Schedule: 04:00 UTC daily  ·  9 tasks  ·  4 task groups  ·  SLA: Gold ready by 08:00 SAST"
ws6.cell(3, 2).font = font(color=MUTED, size=9, italic=True)
ws6.cell(3, 2).fill = fill(DARK)

header_row(ws6, 5, ["","Task","Task Group","Operator Type","Depends On","Est. Duration"], PURPLE)
dag_tasks = [
    ("","trigger_ga4_transfer",        "ingest","BQ Data Transfer Operator",  "—",                 "~2 min"),
    ("","trigger_google_ads_transfer",  "ingest","BQ Data Transfer Operator",  "—",                 "~2 min"),
    ("","trigger_cm360_transfer",       "ingest","BQ Data Transfer Operator",  "—",                 "~2 min"),
    ("","trigger_dv360_transfer",       "ingest","BQ Data Transfer Operator",  "—",                 "~3 min"),
    ("","trigger_youtube_transfer",     "ingest","BQ Data Transfer Operator",  "—",                 "~2 min"),
    ("","run_beam_pipeline",        "validate","PythonOperator (DataflowHook)","ingest (all 5)",    "~8 min"),
    ("","dbt_run_staging_models",  "transform","BashOperator (dbt run)",       "validate",          "~4 min"),
    ("","dbt_test_staging_models", "transform","BashOperator (dbt test)",      "dbt_run_staging",   "~2 min"),
    ("","build_gold_fact_table",       "gold",  "BashOperator (dbt run)",      "transform",         "~3 min"),
]
group_accents = {
    "ingest":    AMBER,
    "validate":  RED,
    "transform": TEAL,
    "gold":      GREEN,
}
for i, row in enumerate(dag_tasks):
    r = 6 + i
    data_row(ws6, r, row, alt=(i%2==1))
    grp = row[2]
    ws6.cell(r, 3).font = font(bold=True, color=group_accents.get(grp, MUTED), size=9)

# Totals row
total_row(ws6, 15, ["","TOTAL","4 task groups","9 tasks","Sequential + parallel","~28 min"])

ws6.row_dimensions[17].height = 16
ws6.cell(17, 2).value = "DESIGN PRINCIPLE: Gold layer is fully overwritten (not appended) on each run — ensures idempotency and correct handling of late-arriving DV360 billing adjustments."
ws6.cell(17, 2).font = font(color=AMBER, size=8, italic=True)
ws6.cell(17, 2).fill = fill(DARK)

# Task group bar (approximate durations)
for i, (grp, mins) in enumerate([("ingest",3),("validate",8),("transform",6),("gold",3)]):
    ws6.cell(21+i, 2, grp)
    ws6.cell(21+i, 3, mins)
bc3 = BarChart()
bc3.type = "bar"
bc3.title = "Approx Duration per Task Group (minutes)"
bc3.y_axis.title = "Minutes"
bc3.add_data(Reference(ws6, min_col=3, min_row=21, max_row=24))
bc3.set_categories(Reference(ws6, min_col=2, min_row=21, max_row=24))
bc3.series[0].graphicalProperties.solidFill = PURPLE
bc3.width = 14; bc3.height = 9
ws6.add_chart(bc3, "E5")

# ────────────────────────────────────────────────────────────────────────────
# SAVE
# ────────────────────────────────────────────────────────────────────────────
import os, pathlib
from datetime import datetime
base_dir = pathlib.Path(__file__).parent
out_path = base_dir / "GCP_Marketing_Analytics_Platform_v2.xlsx"
try:
    wb.save(out_path)
except PermissionError:
    out_path = base_dir / "GCP_Marketing_Analytics_Platform_v2_revitalized.xlsx"
    try:
        wb.save(out_path)
        print("Original workbook was locked; saved refreshed copy instead.")
    except PermissionError:
        stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        out_path = base_dir / f"GCP_Marketing_Analytics_Platform_v2_revitalized_{stamp}.xlsx"
        wb.save(out_path)
        print("Original and refreshed workbooks were locked; saved timestamped copy instead.")
print(f"Saved: {out_path}")
print("Sheets:", [s.title for s in wb.worksheets])
